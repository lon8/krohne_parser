import requests
import time
import random
from openpyxl import load_workbook, Workbook
from multiprocessing import Pool, Manager
from concurrent.futures import ThreadPoolExecutor
from loguru import logger
from tqdm import tqdm

# Настройка логирования
logger.add("app.log", rotation="10 MB", retention="10 days", level="DEBUG")

def fetch(url, serial):
    try:
        logger.debug(f"Starting request for serial: {serial}")
        time.sleep(random.uniform(1, 3))  # случайная пауза от 1 до 3 секунд

        response = requests.get(url)
        if response.status_code == 200:
            logger.debug(f"Request for serial {serial} succeeded with status {response.status_code}")
            return serial, response.json()
        else:
            logger.warning(f"Request to {url} returned status {response.status_code}")
            return serial, None
    except Exception as e:
        logger.error(f"Request to {url} for serial {serial} failed: {e}")
        return serial, None

def process_serials(serials_chunk, output_list):
    logger.info(f"Processing chunk with {len(serials_chunk)} serials in a new process")
    with ThreadPoolExecutor(max_workers=10) as executor:
        results = list(executor.map(lambda serial: fetch(f'https://pick.krohne.com/api/modern/device?serial={serial}', serial), serials_chunk))
        output_list.extend(results)
    logger.info(f"Finished processing chunk with {len(serials_chunk)} serials")

def main(input_file, output_file):
    logger.info("Starting main process")

    # Загрузка серийных номеров
    wb_input = load_workbook(filename=input_file, read_only=True)
    ws_input = wb_input.active
    serials = [row[0] for row in ws_input.iter_rows(min_col=1, max_col=1, values_only=True) if row[0] is not None and "б/n" not in row[0].lower()]
    logger.info(f"Loaded {len(serials)} valid serials from input file (excluding 'б/n')")

    # Разделение на части для параллельной обработки
    chunks = [serials[i::4] for i in range(4)]

    # Используем менеджер для списков
    with Manager() as manager:
        output_list = manager.list()

        # Пул процессов
        with Pool(processes=4) as pool:
            pool.starmap(process_serials, [(chunk, output_list) for chunk in chunks])

        # Создание нового файла в write_only режиме
        wb_output = Workbook(write_only=True)
        ws_output = wb_output.create_sheet(title="Result")

        # Инициализация переменной для заголовков
        headers = ["Serial"]
        rows_to_write = []

        # Обработка данных и сбор заголовков
        for serial, data in tqdm(output_list, desc="Processing data", unit="row"):
            if data is not None and isinstance(data, dict):  # Проверка, что data является словарем
                row_data = {"Serial": serial}
                for item in data.get("deviceTextStructured", []):
                    if "pairLine" in item:
                        name = item["pairLine"]["name"]
                        value = item["pairLine"]["value"]

                        # Добавляем заголовок, если он ещё не добавлен
                        if name not in headers:
                            headers.append(name)

                        row_data[name] = value

                rows_to_write.append(row_data)
            else:
                logger.warning(f"Unexpected data format for serial {serial}: {data}")

        # Запись заголовков
        ws_output.append(headers)

        # Запись данных в файл с прогрессом
        logger.info("Starting to write results to output file")
        for row_data in tqdm(rows_to_write, desc="Writing to Excel", unit="row"):
            row = [row_data.get(header, None) for header in headers]
            ws_output.append(row)

        wb_output.save(output_file)
        logger.info(f"Finished writing results to {output_file}")

if __name__ == '__main__':
    input_file = 'input_file.xlsx'  # Укажите путь к входному xlsx файлу
    output_file = 'output_file.xlsx'  # Укажите путь к выходному xlsx файлу
    logger.info(f"Program started with input file: {input_file} and output file: {output_file}")
    main(input_file, output_file)
    logger.info("Program finished successfully")
